"""
export/extract_oe4.py — Extractor OE4: Plano de Financiamento do Investimento.

Lê o engine GrestelPy (cenário Base, hub_on=True) e produz um workbook Excel
auto-contido com todos os mapas exigidos pela ficha OE4:
  01_Pre_Projeto       — equilíbrio financeiro pré-projeto (2024)
  02_Mapa_Investimento — CAPEX + FM + reservas por ano
  03_Estrutura_Capital — critérios + rácios CP/CA projetados
  04_Plano_Financ      — resumo das fontes (tipo, montante, taxa, prazo)
  05_SD_BancoHub       — serviço da dívida Banco Hub
  06_SD_BEI            — serviço da dívida Linha BEI
  07_PT2030            — mapa subsídio PT2030
  08_SD_Existente      — serviço dívida pré-existente (agregado + por banco)
  09_Pos_Projeto       — equilíbrio financeiro pós-projeto (AF ≥ 30%)

Uso:
  cd GrestelPy_G18_v0.9.5-main
  python export/extract_oe4.py
"""

from __future__ import annotations

import sys
from pathlib import Path
from datetime import date

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(_ROOT / "src"))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── constantes de estilo ──────────────────────────────────────────────────────

_BLUE_DARK  = "1F4E79"
_BLUE_MID   = "2E75B6"
_BLUE_LIGHT = "D6E4F0"
_GREY_LIGHT = "F2F2F2"
_GREEN_SOFT = "E2EFDA"
_ORANGE     = "FCE4D6"

_HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=9)
_HDR_FILL   = PatternFill("solid", fgColor=_BLUE_DARK)
_SUB_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=9)
_SUB_FILL   = PatternFill("solid", fgColor=_BLUE_MID)
_TITLE_FONT = Font(bold=True, name="Arial", size=11, color=_BLUE_DARK)
_LABEL_FONT = Font(name="Arial", size=9)
_BOLD_FONT  = Font(bold=True, name="Arial", size=9)
_THIN       = Side(style="thin", color="BFBFBF")
_THIN_BORD  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_EUR  = '#,##0.00 €'
_EUR0 = '#,##0 €'
_PCT  = '0.00%'
_PCT1 = '0.0%'
_DEC2 = '#,##0.00'
_INT  = '#,##0'


# ── helpers de escrita ────────────────────────────────────────────────────────

def _set_col_width(ws, col_idx: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def _title(ws, row: int, text: str) -> None:
    c = ws.cell(row=row, column=1, value=text)
    c.font = _TITLE_FONT


def _subtitle(ws, row: int, col: int, text: str) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, name="Arial", size=9, color=_BLUE_MID)


def _header_row(ws, row: int, labels: list[str], fill=None) -> None:
    f = fill or _HDR_FILL
    font = _HDR_FONT
    for col_idx, lbl in enumerate(labels, 1):
        c = ws.cell(row=row, column=col_idx, value=lbl)
        c.font = font
        c.fill = f
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = _THIN_BORD


def _data_cell(ws, row: int, col: int, value, fmt: str | None = None,
               bold: bool = False, fill_color: str | None = None,
               align: str = "right") -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = _BOLD_FONT if bold else _LABEL_FONT
    c.alignment = Alignment(horizontal=align)
    c.border = _THIN_BORD
    if fmt:
        c.number_format = fmt
    if fill_color:
        c.fill = PatternFill("solid", fgColor=fill_color)


def _write_kv_table(ws, start_row: int, start_col: int,
                    rows: list[tuple[str, object, str | None]],
                    label_width: int = 36, value_width: int = 18) -> int:
    """Tabela chave-valor simples. Devolve próxima linha disponível."""
    for label, value, fmt in rows:
        lc = ws.cell(row=start_row, column=start_col, value=label)
        lc.font = _LABEL_FONT
        lc.border = _THIN_BORD
        vc = ws.cell(row=start_row, column=start_col + 1, value=value)
        vc.font = _LABEL_FONT
        vc.alignment = Alignment(horizontal="right")
        vc.border = _THIN_BORD
        if fmt:
            vc.number_format = fmt
        start_row += 1
    ws.column_dimensions[get_column_letter(start_col)].width = label_width
    ws.column_dimensions[get_column_letter(start_col + 1)].width = value_width
    return start_row


def _write_df_section(ws, start_row: int, title: str, df: pd.DataFrame,
                      col_fmts: dict[str, str] | None = None,
                      col_widths: dict[str, float] | None = None,
                      totals_row: bool = False) -> int:
    """Escreve título + cabeçalho + dados. Devolve próxima linha livre."""
    _title(ws, start_row, title)
    start_row += 1
    _header_row(ws, start_row, list(df.columns))
    start_row += 1

    for r_idx, (_, row) in enumerate(df.iterrows()):
        alt_fill = _GREY_LIGHT if r_idx % 2 == 1 else None
        is_total = totals_row and r_idx == len(df) - 1
        for c_idx, (col, val) in enumerate(zip(df.columns, row), 1):
            fmt = (col_fmts or {}).get(col)
            _data_cell(ws, start_row, c_idx, val if pd.notna(val) else None,
                       fmt=fmt, bold=is_total,
                       fill_color=(_GREEN_SOFT if is_total else alt_fill))
        start_row += 1

    # larguras de coluna
    if col_widths:
        for col_name, w in col_widths.items():
            if col_name in df.columns:
                _set_col_width(ws, df.columns.get_loc(col_name) + 1, w)

    return start_row + 1


def _meta_footer(ws, row: int) -> None:
    c = ws.cell(row=row, column=1,
                value=f"Gerado em {date.today().isoformat()} | GrestelPy v0.9.5 | Cenário Base + Hub")
    c.font = Font(italic=True, name="Arial", size=8, color="808080")


# ── carregamento do modelo ───────────────────────────────────────────────────

def _load_all():
    from engine.modelo.model import run_model
    from engine.inputs import load as eng_load
    from engine.investimento.investimento import investimento_anual
    from engine.financiamento.financiamento import financiamento_anual
    from engine.projetos.hub_logistico import hub_financing, hub_capex

    print("A carregar o modelo (hub_on=True)…")
    dfs = run_model(cenario="Base", hub_on=True, ecogres_on=False)
    a, base, sched = eng_load(cenario="Base")

    df_inv  = investimento_anual(a, base, sched)
    df_fin  = financiamento_anual(sched, a)
    hub_raw = a.raw.get("hub_logistico", {})
    df_hub_fin  = hub_financing(hub_raw)
    df_hub_capex = hub_capex(hub_raw)
    print("Modelo carregado.")
    return dfs, a, base, sched, df_inv, df_fin, df_hub_fin, df_hub_capex


# ── Sheet 01 — Equilíbrio Pré-Projeto ────────────────────────────────────────

def sheet_pre_projeto(wb: Workbook, base, dfs: dict) -> None:
    ws = wb.create_sheet("01_Pre_Projeto")
    ws.freeze_panes = "A4"

    _title(ws, 1, "01 | Situação de Equilíbrio Financeiro Pré-Projeto (2024)")
    ws.cell(row=2, column=1, value="Fonte: R&C 2024 auditado — Grestel, Produtos Cerâmicos S.A.").font = \
        Font(italic=True, name="Arial", size=8, color="808080")

    # --- helpers de soma
    anc_d  = base.balanco["ativo_nao_corrente"]
    ac_d   = base.balanco["ativo_corrente"]
    cp_d   = base.balanco["capital_proprio"]
    pas_d  = base.balanco["passivo"]

    anc_total = sum(float(v) for v in anc_d.values())
    ac_total  = sum(float(v) for v in ac_d.values())
    atl_total = anc_total + ac_total

    cp_total  = sum(float(v) for v in cp_d.values())
    pass_total = sum(float(v) for v in pas_d.values())

    emp_mlp    = float(pas_d.get("Emprestimos_NC", 0))
    emp_cp_val = float(pas_d.get("Emprestimos_C",  0))
    caixa_val  = float(ac_d.get("Caixa", 0))

    outros_anc = anc_total - float(anc_d["AFT_liquido"]) - float(anc_d["Goodwill"]) - float(anc_d["Subsidiarias"])

    # --- Balanço resumido
    _subtitle(ws, 4, 1, "Balanço Resumido — 31/12/2024")
    kv_balanco = [
        ("Ativo Não Corrente (€)",         anc_total,                        _EUR0),
        ("  — AFT líquido",                float(anc_d["AFT_liquido"]),      _EUR0),
        ("  — Goodwill",                   float(anc_d["Goodwill"]),         _EUR0),
        ("  — Investimentos Subsidiárias", float(anc_d["Subsidiarias"]),     _EUR0),
        ("  — Outros Ativos NC",           outros_anc,                       _EUR0),
        ("Ativo Corrente (€)",             ac_total,                         _EUR0),
        ("  — Inventários",                float(ac_d["Inventarios"]),       _EUR0),
        ("  — Clientes",                   float(ac_d["Clientes"]),          _EUR0),
        ("  — Caixa e equivalentes",       float(ac_d["Caixa"]),             _EUR0),
        ("Ativo Total Líquido (€)",        atl_total,                        _EUR0),
        ("", None, None),
        ("Capital Próprio (€)",            cp_total,                         _EUR0),
        ("  — Capital social",             float(cp_d.get("Capital_Social", 0)),      _EUR0),
        ("  — Outros instrumentos CP",     float(cp_d.get("Outros_IC_Proprio", 0)),   _EUR0),
        ("  — Reservas e RL transitados",  float(cp_d.get("Reservas_Legais", 0))
                                           + float(cp_d.get("Resultados_Transitados", 0))
                                           + float(cp_d.get("Ajust_AF", 0))
                                           + float(cp_d.get("Outras_Var_CP", 0)),     _EUR0),
        ("  — Resultado líquido 2024",     float(cp_d.get("RL_2024", 0)),             _EUR0),
        ("Passivo (total, €)",             pass_total,                               _EUR0),
        ("  — Empréstimos MLP (NC)",       emp_mlp,                                  _EUR0),
        ("  — Empréstimos CP",             emp_cp_val,                               _EUR0),
        ("  — Fornecedores",               float(pas_d.get("Fornecedores", 0)),      _EUR0),
        ("  — Outros passivos",            float(pas_d.get("Outros_PC", 0)) + float(pas_d.get("Impostos_Diferidos_Passivos", 0)), _EUR0),
        ("Total CP + Passivo (€)",         cp_total + pass_total,                    _EUR0),
    ]
    next_row = _write_kv_table(ws, 5, 1, kv_balanco)

    # --- Rácios
    row = next_row + 1
    _subtitle(ws, row, 1, "Rácios de Equilíbrio Financeiro")
    row += 1

    cp      = cp_total
    atl     = atl_total
    passivo = pass_total
    div_fin = emp_mlp + emp_cp_val
    div_liq    = div_fin - caixa_val

    # EBITDA 2024 a partir da DR no dfs
    dr = dfs.get("dr", pd.DataFrame())
    row_2024 = dr[dr["ano"] == 2024] if not dr.empty and "ano" in dr.columns else pd.DataFrame()
    ebitda_2024 = float(row_2024["ebitda"].iloc[0]) if not row_2024.empty and "ebitda" in row_2024.columns else 4_149_679.0
    ebit_2024   = float(row_2024["ebit"].iloc[0])   if not row_2024.empty and "ebit" in row_2024.columns  else 1_980_964.0
    juros_2024  = abs(float(row_2024["juros"].iloc[0])) if not row_2024.empty and "juros" in row_2024.columns else 528_161.0

    af   = cp / atl
    solv = cp / passivo if passivo else 0
    end  = passivo / atl
    gear = div_liq / (div_liq + cp) if (div_liq + cp) else 0
    cob_j = ebit_2024 / juros_2024 if juros_2024 else 0
    dnd_ebitda = div_liq / ebitda_2024 if ebitda_2024 else 0

    kv_ratios = [
        ("Autonomia Financeira  [CP / ATL]",              af,           _PCT),
        ("Solvabilidade  [CP / Passivo]",                 solv,         _PCT),
        ("Endividamento  [Passivo / ATL]",                end,          _PCT),
        ("Gearing  [Dívida Líq. / (Dív.Líq.+CP)]",       gear,         _PCT),
        ("Dívida Financeira Total (€)",                   div_fin,      _EUR0),
        ("Dívida Líquida (€)",                            div_liq,      _EUR0),
        ("EBITDA 2024 (€)",                               ebitda_2024,  _EUR0),
        ("Dívida Líquida / EBITDA  [x]",                  dnd_ebitda,   _DEC2),
        ("Cobertura de Juros  [EBIT / Juros]",            cob_j,        _DEC2),
        ("", None, None),
        ("Condição AF mínima exigida (OE4)",               0.30,         _PCT),
        ("Condição satisfeita pré-projeto?",
         "SIM ✓" if af >= 0.30 else "NÃO ✗", None),
    ]
    row = _write_kv_table(ws, row, 1, kv_ratios)

    _meta_footer(ws, row + 2)
    _set_col_width(ws, 1, 40)
    _set_col_width(ws, 2, 20)


# ── Sheet 02 — Mapa de Investimento ──────────────────────────────────────────

def sheet_mapa_investimento(wb: Workbook, a, base, sched,
                            df_hub_capex: pd.DataFrame) -> None:
    ws = wb.create_sheet("02_Mapa_Investimento")

    _title(ws, 1, "02 | Mapa de Investimento do Projeto")
    ws.cell(row=2, column=1,
            value="Inclui: CAPEX Hub (AFT), CAPEX Core Grestel, Investimento em FM e Reservas de Segurança").font = \
        Font(italic=True, name="Arial", size=8, color="808080")

    hub = a.raw.get("hub_logistico", {})
    proj = hub.get("projeto_hub", {})

    # --- 2a. CAPEX Hub por ano (do módulo hub_capex)
    hub_capex_rows = []
    for _, r in df_hub_capex.iterrows():
        hub_capex_rows.append({
            "Ano":       int(r["ano"]),
            "CAPEX Hub AFT (€)": float(r.get("capex", 0)),
            "Juros Capitalizados (€)": float(r.get("juros_capitalizados_aft", 0)),
            "Depreciação Hub (€)": float(r.get("depreciacao", 0)),
            "AFT Hub Líquido Fim Ano (€)": float(r.get("aft_liquido_fim", 0)),
        })
    df_hub_capex_show = pd.DataFrame(hub_capex_rows)
    fmts_hub = {c: _EUR0 for c in df_hub_capex_show.columns if c != "Ano"}
    fmts_hub["Ano"] = _INT
    r = _write_df_section(ws, 4, "2.1 — CAPEX Hub Logístico (por ano, €)", df_hub_capex_show, fmts_hub)

    # --- 2b. CAPEX Core Grestel
    inv = sched.investimento
    core_aft   = inv.get("novo_investimento_aft", {})
    core_intg  = inv.get("novo_investimento_intang", {})
    from engine.inputs.constants import ALL_YEARS
    core_rows = []
    for y in ALL_YEARS:
        if y >= 2025:
            core_rows.append({
                "Ano": y,
                "CAPEX AFT Core (€)": float(core_aft.get(y, 0)),
                "CAPEX Intangíveis Core (€)": float(core_intg.get(y, 0)),
                "Total CAPEX Core (€)": float(core_aft.get(y, 0)) + float(core_intg.get(y, 0)),
            })
    df_core = pd.DataFrame(core_rows)
    fmts_core = {c: _EUR0 for c in df_core.columns if c != "Ano"}
    fmts_core["Ano"] = _INT
    r = _write_df_section(ws, r, "2.2 — CAPEX Core Grestel (por ano, €)", df_core, fmts_core)

    # --- 2c. Investimento em FM (NFM incremental Hub)
    nfm_raw = proj.get("necessidades_fundo_maneio", {})
    fm_rows = []
    total_fm = 0.0
    for y in sorted(k for k in nfm_raw if isinstance(k, int)):
        v = float(nfm_raw[y])
        total_fm += v
        fm_rows.append({"Ano": y, "Δ NFM Hub (€)": v, "NFM Acumulado (€)": total_fm})
    if fm_rows:
        df_fm = pd.DataFrame(fm_rows)
        fmts_fm = {"Ano": _INT, "Δ NFM Hub (€)": _EUR0, "NFM Acumulado (€)": _EUR0}
        r = _write_df_section(ws, r, "2.3 — Investimento em Fundo de Maneio (Hub)", df_fm, fmts_fm)

    # --- 2d. Resumo total
    hub_total_capex = sum(float(row.get("capex", 0)) for _, row in df_hub_capex.iterrows())
    core_total_aft  = sum(float(core_aft.get(y, 0)) for y in range(2025, 2030))
    core_total_intg = sum(float(core_intg.get(y, 0)) for y in range(2025, 2030))
    pt2030_montante = float(proj.get("financiamento", {}).get("PT2030", {}).get("montante", 0))

    summary = [
        ("Rubrica",                         "Valor (€)",         None),
    ]
    _subtitle(ws, r, 1, "2.4 — Resumo do Investimento Total 2025-2029")
    r += 1
    kv_inv = [
        ("CAPEX Hub — Construção e Equipamento (€)",   hub_total_capex,               _EUR0),
        ("CAPEX Core Grestel — AFT (€)",               core_total_aft,                _EUR0),
        ("CAPEX Core Grestel — Intangíveis (€)",       core_total_intg,               _EUR0),
        ("Investimento em Fundo de Maneio (€)",        total_fm,                      _EUR0),
        ("Reservas de Segurança (mínimo operacional)", 0.0,                           _EUR0),
        ("TOTAL INVESTIMENTO (€)",                     hub_total_capex + core_total_aft + core_total_intg + total_fm, _EUR0),
        ("", None, None),
        ("  dos quais financiados por PT2030 (subsídio)", pt2030_montante,             _EUR0),
        ("  dos quais financiados por capitais alheios (bancário)", 0.0,              _EUR0),
    ]
    r = _write_kv_table(ws, r, 1, kv_inv)
    _meta_footer(ws, r + 1)
    _set_col_width(ws, 1, 46)
    _set_col_width(ws, 2, 22)


# ── Sheet 03 — Estrutura de Capital ──────────────────────────────────────────

def sheet_estrutura_capital(wb: Workbook, dfs: dict) -> None:
    ws = wb.create_sheet("03_Estrutura_Capital")

    _title(ws, 1, "03 | Critérios e Estrutura de Capital do Projeto")
    ws.cell(row=2, column=1,
            value="Projeção 2025-2029: Cenário Base com Hub Logístico").font = \
        Font(italic=True, name="Arial", size=8, color="808080")

    # Critérios
    _subtitle(ws, 4, 1, "3.1 — Critérios para a Estrutura de Capital")
    criterios = [
        ("Critério",                                        "Orientação / Decisão"),
        ("Manutenção da Autonomia Financeira",              "AF ≥ 30% em todos os anos (covenant bancário BPI e condição OE4)"),
        ("Alavancagem moderada",                            "Dívida Líq./EBITDA ≤ 4,0x — limita o risco de refinanciamento"),
        ("Custo ponderado de capital (WACC)",               "Minimizar custo com diversificação de fontes: banco, BEI, PT2030"),
        ("Preservação de CP para autofinanciamento",        "RL retido cobre CAPEX Core anual; evitar aumentos de capital"),
        ("Alívio fiscal",                                   "Juros dedutíveis em IRC (art. 23.º CIRC); RFAI para equipamentos elegíveis"),
        ("Elegibilidade PT2030 (COMPETE 2030 — SI Inovação)", "Candidatura a fundo perdido; reduz CA e melhora AF no ano de recebimento"),
        ("Financiamento BEI via banco parceiro",            "Taxa subsidiada (~3,75%); garante diversificação sem exposição cambial"),
    ]
    r = 5
    _header_row(ws, r, ["Critério", "Orientação / Decisão"])
    r += 1
    for i, (crit, orient) in enumerate(criterios):
        fill = _GREY_LIGHT if i % 2 == 1 else None
        _data_cell(ws, r, 1, crit,  bold=True,  fill_color=fill, align="left")
        _data_cell(ws, r, 2, orient, fill_color=fill, align="left")
        r += 1

    # Projeção rácios
    r += 1
    _subtitle(ws, r, 1, "3.2 — Projeção dos Rácios de Equilíbrio Financeiro (2024-2029)")
    r += 1

    kpis = dfs.get("kpis", pd.DataFrame())
    if not kpis.empty:
        cols_show = [c for c in [
            "ano", "autonomia_financeira", "solvabilidade", "endividamento",
            "debt_ebitda", "nd_ebitda", "cobertura_juros", "dscr",
        ] if c in kpis.columns]
        df_k = kpis[cols_show].copy()
        label_map = {
            "ano": "Ano",
            "autonomia_financeira": "Autonomia Fin. [%]",
            "solvabilidade": "Solvabilidade [%]",
            "endividamento": "Endividamento [%]",
            "debt_ebitda": "Dívida/EBITDA [x]",
            "nd_ebitda": "DívLíq/EBITDA [x]",
            "cobertura_juros": "Cobertura Juros [x]",
            "dscr": "DSCR [x]",
        }
        df_k.columns = [label_map.get(c, c) for c in df_k.columns]
        fmts_k: dict[str, str] = {}
        for col in df_k.columns:
            if col == "Ano":
                fmts_k[col] = _INT
            elif "%" in col:
                fmts_k[col] = _PCT1
            else:
                fmts_k[col] = _DEC2
        r = _write_df_section(ws, r, "", df_k, fmts_k)

    # Referência de covenants
    r += 1
    _subtitle(ws, r, 1, "3.3 — Referências e Covenants")
    r += 1
    kv_cov = [
        ("AF mínima exigida (OE4 e covenant BPI)", 0.30,  _PCT),
        ("AF mínima recomendada (orientação prudencial)", 0.35, _PCT),
        ("DSCR mínimo aceitável", 1.20, _DEC2),
        ("Dívida Líquida / EBITDA máx. (orientação)", 4.0, _DEC2),
    ]
    r = _write_kv_table(ws, r, 1, kv_cov)

    _meta_footer(ws, r + 1)
    _set_col_width(ws, 1, 48)
    _set_col_width(ws, 2, 56)


# ── Sheet 04 — Plano de Financiamento (resumo) ───────────────────────────────

def sheet_plano_financiamento(wb: Workbook, a) -> None:
    ws = wb.create_sheet("04_Plano_Financ")

    _title(ws, 1, "04 | Plano de Financiamento do Investimento")
    ws.cell(row=2, column=1,
            value="Mínimo 2 fontes de capital alheio distintas + subsídio PT2030").font = \
        Font(italic=True, name="Arial", size=8, color="808080")

    hub = a.raw["hub_logistico"]["projeto_hub"]
    fin_hub = hub["financiamento"]

    bh  = fin_hub["Banco_Hub"]
    bei = fin_hub["Linha_BEI"]
    pt  = fin_hub["PT2030"]

    headers = ["Fonte", "Tipo", "Montante (€)", "Taxa Juro (a.a.)", "Prazo (anos)",
               "Início Amort.", "Amort. Anual (€)", "Desembolso", "Reembolsável?"]
    _header_row(ws, 4, headers)

    def _row(ws, r, vals, fill=None):
        for c_idx, v in enumerate(vals, 1):
            fmt = None
            if isinstance(v, float) and abs(v) > 1:
                fmt = _EUR0
            elif isinstance(v, float) and abs(v) < 1:
                fmt = _PCT
            _data_cell(ws, r, c_idx, v, fmt=fmt, fill_color=fill, align="left")

    fontes = [
        ("Banco Hub (BPI/CA)", "Empréstimo Bancário",
         float(bh["montante"]), float(bh["taxa_juro"]),
         10, int(bh["inicio_amortizacao"]),
         float(bh["amortizacao_anual"]), int(bh["desembolso"]), "Sim"),
        ("Linha BEI (via banco parceiro)", "Empréstimo BEI",
         float(bei["montante"]), float(bei["taxa_juro"]),
         10, int(bei["inicio_amortizacao"]),
         float(bei["amortizacao_anual"]), int(bei["desembolso"]), "Sim"),
        ("PT2030 — COMPETE 2030 SI Inovação", "Subsídio Não Reembolsável",
         float(pt["montante"]), 0.0,
         0, int(pt.get("ano_recebimento", 2027)),
         0.0, int(pt.get("ano_recebimento", 2027)), "Não"),
    ]

    fills = [None, _GREY_LIGHT, None]
    r = 5
    for frow, fill in zip(fontes, fills):
        _row(ws, r, frow, fill=fill)
        r += 1

    # Totais
    total_ca = float(bh["montante"]) + float(bei["montante"]) + float(pt["montante"])
    total_row = ["TOTAL FINANCIAMENTO PROJETO", "", total_ca, "", "", "", "", "", ""]
    for c_idx, v in enumerate(total_row, 1):
        c = ws.cell(row=r, column=c_idx, value=v)
        c.font = Font(bold=True, name="Arial", size=9)
        c.fill = PatternFill("solid", fgColor=_BLUE_LIGHT)
        c.border = _THIN_BORD
        if isinstance(v, float) and abs(v) > 1:
            c.number_format = _EUR0
    r += 2

    # Nota sobre CP
    _subtitle(ws, r, 1, "Nota sobre Capitais Próprios")
    r += 1
    notas = [
        "A estrutura de capital não requer reforço de CP no cenário base: AF pré-projeto (30,3%) cumpre",
        "o covenant mínimo de 30% e o plano de financiamento mantém AF ≥ 30% em todos os anos.",
        "Caso a autonomia financeira deça abaixo de 30% nalgum ano, equacionar:",
        "  (a) Suprimento de sócios (reembolsável, custo reduzido, sem diluição imediata);",
        "  (b) Aumento de capital social por incorporação de reservas livres (simples e sem encaixe externo).",
    ]
    for nota in notas:
        c = ws.cell(row=r, column=1, value=nota)
        c.font = Font(italic=True, name="Arial", size=8)
        r += 1

    _meta_footer(ws, r + 1)
    widths = [32, 26, 16, 14, 10, 12, 18, 12, 14]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)


# ── helpers serviço da dívida ─────────────────────────────────────────────────

def _build_amort_table(nome: str, capital: float, taxa: float,
                       amort_anual: float, inicio_amort: int,
                       desembolso: int, years: list[int]) -> pd.DataFrame:
    """Constrói plano de amortização linear."""
    rows = []
    saldo = 0.0
    for y in years:
        if y == desembolso:
            saldo = capital
        juros = saldo * taxa
        amort = amort_anual if y >= inicio_amort and saldo > 0 else 0.0
        amort = min(amort, saldo)
        saldo_ini = saldo
        saldo = max(saldo - amort, 0.0)
        servico = juros + amort
        rows.append({
            "Ano": y,
            "Saldo Inicial (€)": saldo_ini,
            "Amortização Capital (€)": amort,
            "Juros (€)": juros,
            "Serviço Total (€)": servico,
            "Saldo Final (€)": saldo,
        })
    return pd.DataFrame(rows)


def _write_sd_sheet(wb: Workbook, sheet_name: str, title: str,
                    nome_fonte: str, capital: float, taxa: float,
                    amort_anual: float, inicio_amort: int,
                    desembolso: int, tipo: str = "Empréstimo Bancário",
                    nota: str = "") -> None:
    from engine.inputs.constants import YEARS
    ws = wb.create_sheet(sheet_name)
    _title(ws, 1, title)

    # pressupostos
    r = 3
    _subtitle(ws, r, 1, "Pressupostos")
    r += 1
    ps = [
        ("Fonte / Designação",   nome_fonte,  None),
        ("Tipo",                 tipo,         None),
        ("Capital (€)",          capital,      _EUR0),
        ("Taxa de Juro (a.a.)",  taxa,         _PCT),
        ("Amortização Anual (€)", amort_anual, _EUR0),
        ("Início Amortização",   str(inicio_amort), None),
        ("Desembolso",           str(desembolso),   None),
        ("Período de Carência",  f"{inicio_amort - desembolso} ano(s)", None),
    ]
    r = _write_kv_table(ws, r, 1, ps, label_width=30, value_width=22)

    # tabela amortização
    r += 1
    df_sd = _build_amort_table(nome_fonte, capital, taxa, amort_anual,
                               inicio_amort, desembolso, list(YEARS))
    fmts_sd = {"Ano": _INT}
    for col in df_sd.columns:
        if col != "Ano":
            fmts_sd[col] = _EUR0
    r = _write_df_section(ws, r, "Plano de Amortização", df_sd, fmts_sd,
                          totals_row=False)

    # Totais agregados
    _subtitle(ws, r, 1, "Totais do Período")
    r += 1
    kv_tot = [
        ("Total Juros Pagos (€)",           df_sd["Juros (€)"].sum(),              _EUR0),
        ("Total Amortizações Capital (€)",  df_sd["Amortização Capital (€)"].sum(),_EUR0),
        ("Total Serviço da Dívida (€)",     df_sd["Serviço Total (€)"].sum(),      _EUR0),
        ("Capital Reembolsado / Montante",  df_sd["Amortização Capital (€)"].sum() / capital if capital else 0, _PCT),
    ]
    r = _write_kv_table(ws, r, 1, kv_tot, label_width=36, value_width=20)

    if nota:
        r += 1
        c = ws.cell(row=r, column=1, value=nota)
        c.font = Font(italic=True, name="Arial", size=8, color="606060")

    _meta_footer(ws, r + 2)
    _set_col_width(ws, 1, 30)
    _set_col_width(ws, 2, 22)
    for i in range(2, 8):
        _set_col_width(ws, i, 22)


# ── Sheet 05-06 — Serviço Dívida Hub ─────────────────────────────────────────

def sheet_sd_banco_hub(wb: Workbook, a) -> None:
    fin = a.raw["hub_logistico"]["projeto_hub"]["financiamento"]["Banco_Hub"]
    _write_sd_sheet(
        wb, "05_SD_BancoHub",
        "05 | Serviço da Dívida — Banco Hub (Empréstimo Bancário)",
        nome_fonte="Banco Hub (BPI / Banco Comercial Parceiro)",
        capital=float(fin["montante"]),
        taxa=float(fin["taxa_juro"]),
        amort_anual=float(fin["amortizacao_anual"]),
        inicio_amort=int(fin["inicio_amortizacao"]),
        desembolso=int(fin["desembolso"]),
        tipo="Empréstimo Bancário (Euribor 3M + Spread)",
        nota="Taxa indexada: Euribor 3M (2,90%) + Spread 1,25% = 4,15%. Revisão anual."
             " Amortização linear. Carência 2 anos (2025-2026, período de obra + ramp-up).",
    )


def sheet_sd_bei(wb: Workbook, a) -> None:
    fin = a.raw["hub_logistico"]["projeto_hub"]["financiamento"]["Linha_BEI"]
    _write_sd_sheet(
        wb, "06_SD_BEI",
        "06 | Serviço da Dívida — Linha BEI (European Investment Bank)",
        nome_fonte="Linha BEI — EIB Global Loan (via banco parceiro)",
        capital=float(fin["montante"]),
        taxa=float(fin["taxa_juro"]),
        amort_anual=float(fin["amortizacao_anual"]),
        inicio_amort=int(fin["inicio_amortizacao"]),
        desembolso=int(fin["desembolso"]),
        tipo="Empréstimo BEI (taxa fixa subsidiada)",
        nota="Taxa fixa 3,75% (benefício BEI vs. mercado ~4,15%). Desembolso 2026 após conclusão da 1.ª fase."
             " Elegível por se tratar de investimento em infraestrutura logística e digitalização (NCEF / InvestEU).",
    )


# ── Sheet 07 — PT2030 ────────────────────────────────────────────────────────

def sheet_pt2030(wb: Workbook, a) -> None:
    ws = wb.create_sheet("07_PT2030")
    fin = a.raw["hub_logistico"]["projeto_hub"]["financiamento"]["PT2030"]
    montante = float(fin["montante"])
    ano_rec  = int(fin.get("ano_recebimento", 2027))

    _title(ws, 1, "07 | Subsídio PT2030 — COMPETE 2030 SI Inovação")
    ws.cell(row=2, column=1,
            value="Instrumento não reembolsável — sem serviço de dívida associado").font = \
        Font(italic=True, name="Arial", size=8, color="808080")

    r = 4
    _subtitle(ws, r, 1, "Caracterização do Instrumento")
    r += 1
    kv = [
        ("Programa",                     "Portugal 2030 — COMPETE 2030",               None),
        ("Tipologia",                     "SI Inovação Produtiva (Aviso n.º COMPETE-SI)", None),
        ("Natureza",                      "Subsídio não reembolsável (fundo perdido)",   None),
        ("Montante Previsto (€)",         montante,                                      _EUR0),
        ("Ano de Recebimento Previsto",   str(ano_rec),                                  None),
        ("Critérios de Elegibilidade",
         "Investimento produtivo, PME ou Mid-Cap, atividade CAE 23 (cerâmica), região Centro", None),
        ("Contrapartida em CP Próprio",
         "Mínimo 25% do investimento elegível financiado por capitais próprios ou equivalente", None),
        ("Impacto no Balanço",
         "Reconhecido como Subsídio ao Investimento (rendimento diferido, amortizado com o ativo)", None),
        ("Tratamento Fiscal",
         "Rendimento tributável distribuído ao longo da vida útil do ativo (CIRC art. 22.º)", None),
        ("", None, None),
        ("Impacto na Autonomia Financeira",
         "Reduz capital alheio líquido; melhora AF no ano de receção", None),
        ("Risco principal",
         "Demora ou indeferimento da candidatura → custo de oportunidade mitigado por linha BEI", None),
    ]
    r = _write_kv_table(ws, r, 1, kv, label_width=36, value_width=56)

    r += 1
    _subtitle(ws, r, 1, "Enquadramento Legal")
    r += 1
    legais = [
        "Regulamento (UE) 2021/1060 — disposições comuns fundos estruturais e de investimento;",
        "Regulamento (UE) 2021/1058 — FEDER (base do COMPETE 2030);",
        "Aviso específico de candidatura SI Inovação a publicar pelo IAPMEI / CCDR-C;",
        "CIRC art. 22.º — tributação de subsídios ao investimento por duodécimos ao longo da vida útil.",
    ]
    for linha in legais:
        c = ws.cell(row=r, column=1, value=linha)
        c.font = Font(italic=True, name="Arial", size=8)
        r += 1

    _meta_footer(ws, r + 1)
    _set_col_width(ws, 1, 40)
    _set_col_width(ws, 2, 60)


# ── Sheet 08 — Dívida Existente (pré-projeto) ────────────────────────────────

def sheet_divida_existente(wb: Workbook, sched, df_fin: pd.DataFrame) -> None:
    ws = wb.create_sheet("08_SD_Existente")

    _title(ws, 1, "08 | Serviço da Dívida Pré-Existente (2024-2029)")
    ws.cell(row=2, column=1,
            value="Fontes: BPI, Santander, CGD, Abanca, IAPMEI, Locações Financeiras — dados schedules.yaml").font = \
        Font(italic=True, name="Arial", size=8, color="808080")

    # Agregado (do financiamento_anual)
    df_agg = df_fin[["ano", "juros_base", "amortizacoes_capital",
                     "capital_divida_total_fim", "emprestimos_NC", "emprestimos_C"]].copy()
    df_agg.columns = ["Ano", "Juros (€)", "Amortiz. Capital (€)",
                      "Capital em Dívida FdA (€)", "Empréstimos NC (€)", "Empréstimos C (€)"]
    df_agg["Serviço Total (€)"] = df_agg["Juros (€)"] + df_agg["Amortiz. Capital (€)"]
    cols_order = ["Ano", "Capital em Dívida FdA (€)", "Amortiz. Capital (€)",
                  "Juros (€)", "Serviço Total (€)", "Empréstimos NC (€)", "Empréstimos C (€)"]
    df_agg = df_agg[cols_order]

    fmts_agg: dict[str, str] = {"Ano": _INT}
    for col in df_agg.columns:
        if col != "Ano":
            fmts_agg[col] = _EUR0

    r = _write_df_section(ws, 4, "8.1 — Serviço da Dívida Agregado (Dívida Pré-Existente)", df_agg, fmts_agg)

    # Por banco (a partir de schedules.yaml)
    raw_fin = sched.raw.get("financiamento", {})
    banks = ["BPI", "Santander", "CGD_COVID", "CGD_OS", "Abanca", "IAPMEI", "Locacoes"]
    bank_labels = {
        "BPI": "BPI",
        "Santander": "Santander",
        "CGD_COVID": "CGD (COVID)",
        "CGD_OS": "CGD (OS)",
        "Abanca": "Abanca",
        "IAPMEI": "IAPMEI",
        "Locacoes": "Locações Financeiras",
    }
    from engine.inputs.constants import ALL_YEARS

    r += 1
    _subtitle(ws, r, 1, "8.2 — Saldo de Capital por Banco (fim de ano)")
    r += 1

    bank_rows_data: list[dict] = []
    for y in ALL_YEARS:
        row_d: dict[str, object] = {"Ano": y}
        for b in banks:
            key = f"{b}_capital_fim_ano"
            val = raw_fin.get(key, {})
            row_d[bank_labels[b]] = float(val.get(y, 0)) if isinstance(val, dict) else 0.0
        bank_rows_data.append(row_d)

    df_banks = pd.DataFrame(bank_rows_data)
    fmts_banks: dict[str, str] = {"Ano": _INT}
    for col in df_banks.columns:
        if col != "Ano":
            fmts_banks[col] = _EUR0
    r = _write_df_section(ws, r, "", df_banks, fmts_banks)

    _meta_footer(ws, r + 1)
    for i in range(1, len(df_banks.columns) + 1):
        _set_col_width(ws, i, 20)
    _set_col_width(ws, 1, 8)


# ── Sheet 09 — Equilíbrio Pós-Projeto ────────────────────────────────────────

def sheet_pos_projeto(wb: Workbook, dfs: dict) -> None:
    ws = wb.create_sheet("09_Pos_Projeto")

    _title(ws, 1, "09 | Situação de Equilíbrio Financeiro Pós-Projeto (2025-2029)")
    ws.cell(row=2, column=1,
            value="Condição OE4: AF ≥ 30% em todos os anos do horizonte").font = \
        Font(italic=True, name="Arial", size=8, color="808080")

    kpis = dfs.get("kpis", pd.DataFrame())
    balanco = dfs.get("balanco", pd.DataFrame())

    if not kpis.empty:
        cols_sel = [c for c in [
            "ano", "total_ativo", "cp",
            "autonomia_financeira", "solvabilidade", "endividamento",
            "divida_liquida", "ebitda", "nd_ebitda",
            "cobertura_juros", "dscr",
        ] if c in kpis.columns]
        df_pos = kpis[cols_sel].copy()
        label_map = {
            "ano": "Ano",
            "total_ativo": "Ativo Total (€)",
            "cp": "Capital Próprio (€)",
            "autonomia_financeira": "AF [%]",
            "solvabilidade": "Solvabilidade [%]",
            "endividamento": "Endividamento [%]",
            "divida_liquida": "Dívida Líquida (€)",
            "ebitda": "EBITDA (€)",
            "nd_ebitda": "DívLíq/EBITDA [x]",
            "cobertura_juros": "Cob. Juros [x]",
            "dscr": "DSCR [x]",
        }
        df_pos.columns = [label_map.get(c, c) for c in df_pos.columns]
        fmts_pos: dict[str, str] = {"Ano": _INT}
        for col in df_pos.columns:
            if col == "Ano":
                fmts_pos[col] = _INT
            elif "€" in col:
                fmts_pos[col] = _EUR0
            elif "%" in col:
                fmts_pos[col] = _PCT1
            else:
                fmts_pos[col] = _DEC2

        r = _write_df_section(ws, 4, "Rácios de Equilíbrio Financeiro Projetados (com Hub)", df_pos, fmts_pos)

        # Sinalização AF ≥ 30%
        r += 1
        _subtitle(ws, r, 1, "Verificação da Condição Mínima: AF ≥ 30%")
        r += 1
        _header_row(ws, r, ["Ano", "AF Projetada", "Condição AF ≥ 30%", "Margem (p.p.)"])
        r += 1
        for _, row_k in kpis[["ano", "autonomia_financeira"]].iterrows():
            y = int(row_k["ano"])
            af_val = float(row_k["autonomia_financeira"])
            cond = "✓ CUMPRE" if af_val >= 0.30 else "✗ NÃO CUMPRE"
            margem = af_val - 0.30
            fill = _GREEN_SOFT if af_val >= 0.30 else _ORANGE
            _data_cell(ws, r, 1, y,     fmt=_INT,  fill_color=fill)
            _data_cell(ws, r, 2, af_val, fmt=_PCT1, fill_color=fill)
            _data_cell(ws, r, 3, cond,              fill_color=fill, align="center")
            _data_cell(ws, r, 4, margem, fmt=_PCT1, fill_color=fill)
            r += 1
    else:
        r = 4
        ws.cell(row=r, column=1, value="Dados kpis não disponíveis — verificar run_model().").font = \
            Font(italic=True, name="Arial", size=9, color="FF0000")

    _meta_footer(ws, r + 2)
    _set_col_width(ws, 1, 10)
    for i in range(2, 12):
        _set_col_width(ws, i, 20)


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    dfs, a, base, sched, df_inv, df_fin, df_hub_fin, df_hub_capex = _load_all()

    wb = Workbook()
    wb.remove(wb.active)  # remove sheet vazia padrão

    print("  Sheet 01 — Pré-Projeto…",    end=" ", flush=True)
    sheet_pre_projeto(wb, base, dfs)
    print("ok")

    print("  Sheet 02 — Investimento…",   end=" ", flush=True)
    sheet_mapa_investimento(wb, a, base, sched, df_hub_capex)
    print("ok")

    print("  Sheet 03 — Estr. Capital…",  end=" ", flush=True)
    sheet_estrutura_capital(wb, dfs)
    print("ok")

    print("  Sheet 04 — Plano Financ…",   end=" ", flush=True)
    sheet_plano_financiamento(wb, a)
    print("ok")

    print("  Sheet 05 — SD Banco Hub…",   end=" ", flush=True)
    sheet_sd_banco_hub(wb, a)
    print("ok")

    print("  Sheet 06 — SD BEI…",         end=" ", flush=True)
    sheet_sd_bei(wb, a)
    print("ok")

    print("  Sheet 07 — PT2030…",         end=" ", flush=True)
    sheet_pt2030(wb, a)
    print("ok")

    print("  Sheet 08 — Dívida Exist…",   end=" ", flush=True)
    sheet_divida_existente(wb, sched, df_fin)
    print("ok")

    print("  Sheet 09 — Pós-Projeto…",    end=" ", flush=True)
    sheet_pos_projeto(wb, dfs)
    print("ok")

    out_path = _ROOT / "export" / f"OE04_G18_{date.today().strftime('%Y%m%d')}.xlsx"
    wb.save(str(out_path))
    print(f"\nFicheiro guardado: {out_path}")


if __name__ == "__main__":
    main()
