"""Rota de exportação para Excel."""

import io
from datetime import datetime

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.engine.modelo.model import dataframe_to_records, run_model
from src.engine.modelo.pressupostos import build_pressupostos_summary
from src.engine.projetos.hub_logistico import (
    load as hub_load,
    viabilidade_hub,
    mapa_servico_divida,
)

router = APIRouter(prefix="/api")

_HDR_FILL = PatternFill("solid", fgColor="3B2A1A")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL = PatternFill("solid", fgColor="F7F0E8")
_SEC_FONT = Font(bold=True, color="3B2A1A", size=10)


def _write_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal="left")


def _write_records(ws, records, start_row=2):
    for i, row in enumerate(records):
        fill = _ALT_FILL if i % 2 == 0 else None
        for col, val in enumerate(row.values(), 1):
            cell = ws.cell(row=start_row + i, column=col, value=val)
            if fill:
                cell.fill = fill


def _sheet_from_records(ws, records):
    if not records:
        ws.cell(1, 1, "Sem dados")
        return
    _write_header(ws, list(records[0].keys()))
    _write_records(ws, records)
    _autofit(ws)


def _autofit(ws):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 45)


@router.get("/export/excel")
def export_excel(
    cenario: str = Query("Base"),
    hub_on: bool = Query(False),
    ecogres_on: bool = Query(True),
):
    """Gera ficheiro Excel com todos os dados calculados pelo modelo financeiro."""
    dfs = run_model(cenario=cenario, hub_on=hub_on, ecogres_on=ecogres_on)
    rec = dataframe_to_records(dfs)
    pressupostos = build_pressupostos_summary(cenario=cenario, hub_on=hub_on, ecogres_on=ecogres_on)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # DR
    ws = wb.create_sheet("DR")
    _sheet_from_records(ws, rec.get("dr", []))

    # Balanço
    ws = wb.create_sheet("Balanço")
    _sheet_from_records(ws, rec.get("balanco", []))

    # DFC
    ws = wb.create_sheet("DFC")
    _sheet_from_records(ws, rec.get("dfc", []))

    # KPIs
    ws = wb.create_sheet("KPIs")
    _sheet_from_records(ws, rec.get("kpis", []))

    # FSE
    ws = wb.create_sheet("FSE")
    _sheet_from_records(ws, rec.get("fse_detalhe_anual", []))

    # Pessoal
    ws = wb.create_sheet("Pessoal")
    _sheet_from_records(ws, rec.get("pessoal_contab_anual", []))

    # Produção
    ws = wb.create_sheet("Produção")
    _sheet_from_records(ws, rec.get("producao_anual", []))

    # Pressupostos — flat: Secção | Parâmetro | Valor | Unidade | Nota
    ws = wb.create_sheet("Pressupostos")
    _write_header(ws, ["Secção", "Parâmetro", "Valor", "Unidade", "Nota"])
    row_idx = 2
    for section in pressupostos.get("sections", []):
        sec_label = section.get("label", "")
        for item in section.get("items", []):
            fill = _ALT_FILL if row_idx % 2 == 0 else None
            vals = [sec_label, item.get("label"), item.get("value"), item.get("unit", ""), item.get("note", "")]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                if fill:
                    cell.fill = fill
            row_idx += 1
    _autofit(ws)

    # Hub Logístico (apenas se ativo)
    if hub_on:
        hub = hub_load()
        hub_res = viabilidade_hub(hub)

        # Hub · Viabilidade
        ws_hv = wb.create_sheet("Hub_Viabilidade")
        kv_rows = [
            {"Indicador": "VAL (€)", "Valor": hub_res.get("val")},
            {"Indicador": "TIR", "Valor": hub_res.get("tir")},
            {"Indicador": "Payback Simples (anos)", "Valor": hub_res.get("payback_simples")},
            {"Indicador": "Payback Atualizado (anos)", "Valor": hub_res.get("payback_atualizado")},
            {"Indicador": "Índice de Rendibilidade", "Valor": hub_res.get("indice_rendibilidade")},
            {"Indicador": "Valor Terminal (€)", "Valor": hub_res.get("valor_terminal")},
            {"Indicador": "Valor Residual Ativos (€)", "Valor": hub_res.get("valor_residual_ativos")},
            {"Indicador": "NFM Recovery Terminal (€)", "Valor": hub_res.get("nfm_recovery_terminal")},
            {"Indicador": "Capital Vivo T10 (€)", "Valor": hub_res.get("capital_vivo_t10")},
            {"Indicador": "Mais-Valia (€)", "Valor": hub_res.get("mais_valia")},
        ]
        for k, v in (hub_res.get("parametros") or {}).items():
            kv_rows.append({"Indicador": f"Parâmetro: {k}", "Valor": v})
        _write_header(ws_hv, ["Indicador", "Valor"])
        _write_records(ws_hv, kv_rows)

        # FCF por ano — em bloco abaixo dos indicadores
        fcf_df = hub_res.get("fcf_df")
        if fcf_df is not None and hasattr(fcf_df, "to_dict"):
            fcf_records = fcf_df.to_dict(orient="records")
            if fcf_records:
                start = len(kv_rows) + 3
                ws_hv.cell(start, 1, "Free Cash Flow por Ano").font = _SEC_FONT
                _write_header(ws_hv, list(fcf_records[0].keys()), row=start + 1)
                _write_records(ws_hv, fcf_records, start_row=start + 2)
        _autofit(ws_hv)

        # Hub · Serviço da Dívida
        ws_hd = wb.create_sheet("Hub_Divida")
        debt_df = mapa_servico_divida(hub)
        if hasattr(debt_df, "to_dict"):
            _sheet_from_records(ws_hd, debt_df.to_dict(orient="records"))

    # Info
    ws_i = wb.create_sheet("Info")
    _sheet_from_records(ws_i, [
        {"Parâmetro": "Cenário", "Valor": cenario},
        {"Parâmetro": "Hub Logístico", "Valor": "Ativo" if hub_on else "Desativado"},
        {"Parâmetro": "Ecogres Consolidada", "Valor": "Sim" if ecogres_on else "Não"},
        {"Parâmetro": "Data de exportação", "Valor": datetime.now().strftime("%Y-%m-%d %H:%M")},
        {"Parâmetro": "Modelo", "Valor": "GrestelPy v0.9.5"},
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"GrestelPy_{cenario}{'_Hub' if hub_on else ''}_{timestamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
